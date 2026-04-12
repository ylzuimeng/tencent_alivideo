"""
医生信息管理服务

负责：
- Excel文件解析
- 数据校验
- 批次管理
- 医生信息CRUD
"""
import os
import logging
from datetime import datetime
from utils.time_helpers import utcnow
from typing import Dict, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from models import db, DoctorInfo

logger = logging.getLogger(__name__)


class DoctorService:
    """医生信息管理服务"""

    ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

    # 必填字段
    REQUIRED_FIELDS = ['姓名', '医院', '科室', '职称']

    @staticmethod
    def allowed_file(filename: str) -> bool:
        """检查文件类型是否允许"""
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in DoctorService.ALLOWED_EXTENSIONS

    @staticmethod
    def generate_batch_id() -> str:
        """生成批次号"""
        return utcnow().strftime('%Y%m%d_%H%M%S')

    @staticmethod
    def import_from_excel(file_path: str, batch_id: Optional[str] = None) -> Tuple[bool, str, list]:
        """
        从Excel文件导入医生信息

        Args:
            file_path: Excel文件路径
            batch_id: 批次号，如果不提供则自动生成

        Returns:
            (是否成功, 消息, 导入的数据列表)
        """
        try:
            if not batch_id:
                batch_id = DoctorService.generate_batch_id()

            # 加载工作簿
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # 读取表头
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    break

            logger.info(f"Excel表头: {headers}")

            # 验证必填字段
            missing_fields = [field for field in DoctorService.REQUIRED_FIELDS if field not in headers]
            if missing_fields:
                return False, f"Excel文件缺少必填字段: {', '.join(missing_fields)}", []

            # 创建字段索引映射
            field_indices = {field: headers.index(field) for field in DoctorService.REQUIRED_FIELDS}

            # 读取数据
            doctors = []
            errors = []
            row_num = 1  # 从第2行开始（第1行是表头）

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1

                # 跳过空行
                if not any(row):
                    continue

                try:
                    # 提取字段
                    name = str(row[field_indices['姓名']]).strip() if row[field_indices['姓名']] else ''
                    hospital = str(row[field_indices['医院']]).strip() if row[field_indices['医院']] else ''
                    department = str(row[field_indices['科室']]).strip() if row[field_indices['科室']] else ''
                    title = str(row[field_indices['职称']]).strip() if row[field_indices['职称']] else ''

                    # 验证必填字段
                    if not name:
                        errors.append(f"第{row_num}行: 姓名不能为空")
                        continue

                    if not hospital:
                        errors.append(f"第{row_num}行: 医院不能为空")
                        continue

                    # 创建医生信息对象
                    doctor_data = {
                        'name': name,
                        'hospital': hospital,
                        'department': department,
                        'title': title,
                        'batch_id': batch_id,
                        'is_validated': False
                    }

                    doctors.append(doctor_data)
                    logger.debug(f"解析医生信息: {doctor_data}")

                except Exception as e:
                    errors.append(f"第{row_num}行: {str(e)}")
                    logger.error(f"解析第{row_num}行失败: {str(e)}")

            # 批量保存到数据库
            saved_count = 0
            for doctor_data in doctors:
                try:
                    doctor = DoctorInfo(**doctor_data)
                    db.session.add(doctor)
                    saved_count += 1
                except Exception as e:
                    errors.append(f"保存数据失败: {str(e)}")
                    logger.error(f"保存医生信息失败: {str(e)}")

            # 提交事务
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                return False, f"数据库保存失败: {str(e)}", []

            # 构建结果消息
            message = f"成功导入 {saved_count} 条医生信息"
            if errors:
                message += f"，但有 {len(errors)} 个错误"

            return True, message, doctors

        except InvalidFileException:
            return False, "无效的Excel文件格式", []
        except Exception as e:
            logger.error(f"导入Excel文件失败: {str(e)}")
            return False, f"导入失败: {str(e)}", []

    @staticmethod
    def get_doctors(
        page: int = 1,
        per_page: int = 20,
        search: Optional[str] = None,
        hospital: Optional[str] = None,
        department: Optional[str] = None,
        batch_id: Optional[str] = None
    ) -> tuple:
        """
        获取医生列表（支持分页和筛选）

        Args:
            page: 页码
            per_page: 每页数量
            search: 搜索关键词（姓名）
            hospital: 医院筛选
            department: 科室筛选
            batch_id: 批次号筛选

        Returns:
            (医生列表, 总数)
        """
        query = DoctorInfo.query

        # 应用筛选条件
        if search:
            query = query.filter(DoctorInfo.name.like(f'%{search}%'))
        if hospital:
            query = query.filter(DoctorInfo.hospital.like(f'%{hospital}%'))
        if department:
            query = query.filter(DoctorInfo.department.like(f'%{department}%'))
        if batch_id:
            query = query.filter(DoctorInfo.batch_id == batch_id)

        # 排序
        query = query.order_by(DoctorInfo.created_at.desc())

        # 分页
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        return pagination.items, pagination.total

    @staticmethod
    def get_doctor_by_id(doctor_id: int) -> Optional[DoctorInfo]:
        """根据ID获取医生信息"""
        return DoctorInfo.query.get(doctor_id)

    @staticmethod
    def update_doctor(doctor_id: int, **kwargs) -> Tuple[bool, str]:
        """
        更新医生信息

        Args:
            doctor_id: 医生ID
            **kwargs: 要更新的字段

        Returns:
            (是否成功, 消息)
        """
        try:
            doctor = DoctorInfo.query.get(doctor_id)
            if not doctor:
                return False, "医生信息不存在"

            # 更新字段
            for key, value in kwargs.items():
                if hasattr(doctor, key):
                    setattr(doctor, key, value)

            db.session.commit()
            return True, "更新成功"

        except Exception as e:
            db.session.rollback()
            logger.error(f"更新医生信息失败: {str(e)}")
            return False, f"更新失败: {str(e)}"

    @staticmethod
    def delete_doctor(doctor_id: int) -> Tuple[bool, str]:
        """
        删除医生信息

        Args:
            doctor_id: 医生ID

        Returns:
            (是否成功, 消息)
        """
        try:
            doctor = DoctorInfo.query.get(doctor_id)
            if not doctor:
                return False, "医生信息不存在"

            # 检查是否有关联的处理任务
            if doctor.processing_tasks.count() > 0:
                return False, "该医生信息已关联处理任务，无法删除"

            db.session.delete(doctor)
            db.session.commit()
            return True, "删除成功"

        except Exception as e:
            db.session.rollback()
            logger.error(f"删除医生信息失败: {str(e)}")
            return False, f"删除失败: {str(e)}"

    @staticmethod
    def validate_doctor(doctor_id: int) -> Tuple[bool, str]:
        """
        校验医生信息

        Args:
            doctor_id: 医生ID

        Returns:
            (是否成功, 消息)
        """
        try:
            doctor = DoctorInfo.query.get(doctor_id)
            if not doctor:
                return False, "医生信息不存在"

            # 标记为已校验
            doctor.is_validated = True
            db.session.commit()
            return True, "校验成功"

        except Exception as e:
            db.session.rollback()
            logger.error(f"校验医生信息失败: {str(e)}")
            return False, f"校验失败: {str(e)}"

    @staticmethod
    def get_batch_list() -> list:
        """
        获取所有批次列表

        Returns:
            批次列表
        """
        try:
            # 使用原生SQL查询批次统计
            result = db.session.query(
                DoctorInfo.batch_id,
                db.func.count(DoctorInfo.id).label('count'),
                db.func.max(DoctorInfo.created_at).label('created_at')
            ).group_by(DoctorInfo.batch_id).order_by(
                db.func.max(DoctorInfo.created_at).desc()
            ).all()

            return [
                {
                    'batch_id': r.batch_id,
                    'count': r.count,
                    'created_at': r.created_at.isoformat()
                }
                for r in result
            ]

        except Exception as e:
            logger.error(f"获取批次列表失败: {str(e)}")
            return []

    @staticmethod
    def export_to_excel(batch_id: Optional[str] = None) -> Optional[str]:
        """
        导出医生信息到Excel（功能预留）

        Args:
            batch_id: 批次号，如果指定则只导出该批次

        Returns:
            导出的文件路径
        """
        # TODO: 实现导出功能
        pass


# 全局服务实例
doctor_service = DoctorService()
